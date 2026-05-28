"""
TIME 미국나스닥100 액티브 ETF (426030) — 보유종목 분석기

상장일(2022-05-11)부터 전체 영업일 보유종목을 수집하고,
각 종목의 실제 종가와 매핑하여 포트폴리오를 분석한다.
"""

import os
import json
import time
import urllib.request
import io
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
import yfinance as yf

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"
DATA_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

ETF_IDX = 2
ETF_CATE = "001"
LISTING_DATE = "2022-05-11"
EXCEL_URL = "https://timeetf.co.kr/pdf_excel.php?idx={idx}&cate={cate}&pdfDate={date}"

TICKER_CACHE_PATH = DATA_DIR / "ticker_map.json"
HOLDINGS_CACHE_PATH = DATA_DIR / "all_holdings.csv"
PRICES_CACHE_PATH = DATA_DIR / "all_prices.csv"
SCAN_LOG_PATH = DATA_DIR / "scan_log.json"


def bloomberg_to_yahoo(bbg: str) -> str | None:
    """NVDA US EQUITY → NVDA, TSM US EQUITY → TSM, etc."""
    if not bbg or "Index" in bbg:
        return None
    parts = bbg.strip().split()
    if len(parts) >= 2 and parts[-1] == "EQUITY":
        return parts[0]
    return None


def fetch_holdings_for_date(date_str: str) -> pd.DataFrame | None:
    """단일 날짜의 보유종목을 timeetf.co.kr에서 다운로드."""
    url = EXCEL_URL.format(idx=ETF_IDX, cate=ETF_CATE, date=date_str)
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        resp = urllib.request.urlopen(req, timeout=15)
        data = resp.read()
        if len(data) < 500:
            return None
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        if ws.max_row <= 1:
            return None
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            code, name, qty, eval_amt, weight = row
            ticker = bloomberg_to_yahoo(code) if code else None
            is_cash = name and "현금" in str(name)
            is_futures = code and "Index" in str(code)
            rows.append({
                "date": date_str,
                "bbg_code": code,
                "ticker": ticker,
                "name": name,
                "quantity": qty if not is_cash else 0,
                "eval_krw": int(str(eval_amt).replace(",", "")) if eval_amt else 0,
                "weight_pct": float(weight) if weight else 0.0,
                "is_cash": is_cash,
                "is_futures": is_futures,
            })
        if not rows:
            return None
        return pd.DataFrame(rows)
    except Exception:
        return None


def generate_date_range(start: str, end: str) -> list[str]:
    """start~end 사이의 모든 날짜 (주말 제외)."""
    s = datetime.strptime(start, "%Y-%m-%d")
    e = datetime.strptime(end, "%Y-%m-%d")
    dates = []
    current = s
    while current <= e:
        if current.weekday() < 5:
            dates.append(current.strftime("%Y-%m-%d"))
        current += timedelta(days=1)
    return dates


def scan_all_holdings(start: str = LISTING_DATE, end: str = None,
                      delay: float = 0.3, resume: bool = True) -> pd.DataFrame:
    """상장일~오늘까지 전 영업일 보유종목 스캔. 결과를 parquet으로 캐싱."""
    if end is None:
        end = datetime.now().strftime("%Y-%m-%d")

    scan_log = {}
    if resume and SCAN_LOG_PATH.exists():
        scan_log = json.loads(SCAN_LOG_PATH.read_text())

    all_frames = []
    if resume and HOLDINGS_CACHE_PATH.exists():
        existing = pd.read_csv(HOLDINGS_CACHE_PATH)
        all_frames.append(existing)
        cached_dates = set(existing["date"].unique())
    else:
        cached_dates = set()

    dates = generate_date_range(start, end)
    to_fetch = [d for d in dates if d not in cached_dates and d not in scan_log]
    total = len(to_fetch)

    if total == 0:
        print(f"All {len(dates)} dates already scanned.")
        if all_frames:
            return pd.concat(all_frames, ignore_index=True)
        return pd.DataFrame()

    print(f"Scanning {total} new dates ({start} ~ {end})...")
    new_frames = []
    fetched = 0
    empty = 0

    for i, d in enumerate(to_fetch):
        df = fetch_holdings_for_date(d)
        if df is not None:
            new_frames.append(df)
            fetched += 1
        else:
            empty += 1
            scan_log[d] = "empty"

        if (i + 1) % 50 == 0 or i == total - 1:
            pct = (i + 1) / total * 100
            print(f"  [{i+1}/{total}] {pct:.0f}% — {fetched} with data, {empty} empty")

            SCAN_LOG_PATH.write_text(json.dumps(scan_log, ensure_ascii=False))
            if new_frames:
                combined = pd.concat(all_frames + new_frames, ignore_index=True)
                combined.to_csv(HOLDINGS_CACHE_PATH, index=False)

        time.sleep(delay)

    if new_frames:
        all_df = pd.concat(all_frames + new_frames, ignore_index=True)
    elif all_frames:
        all_df = pd.concat(all_frames, ignore_index=True)
    else:
        all_df = pd.DataFrame()

    if not all_df.empty:
        all_df.to_csv(HOLDINGS_CACHE_PATH, index=False)

    SCAN_LOG_PATH.write_text(json.dumps(scan_log, ensure_ascii=False))
    n_dates = all_df["date"].nunique() if not all_df.empty else 0
    print(f"\nDone. {n_dates} dates with holdings data cached.")
    return all_df


def fetch_all_prices(holdings_df: pd.DataFrame, resume: bool = True) -> pd.DataFrame:
    """보유종목 전체의 가격 히스토리를 yfinance로 일괄 다운로드."""
    if resume and PRICES_CACHE_PATH.exists():
        existing = pd.read_csv(PRICES_CACHE_PATH, index_col=0, parse_dates=True)
        print(f"Loaded cached prices: {len(existing.columns)} tickers, {len(existing)} days")
        new_tickers = set(holdings_df[holdings_df["ticker"].notna()]["ticker"].unique())
        cached_tickers = set(existing.columns)
        missing = new_tickers - cached_tickers
        if not missing:
            return existing
        print(f"Fetching {len(missing)} new tickers...")
        tickers_to_fetch = list(missing)
    else:
        tickers_to_fetch = list(holdings_df[holdings_df["ticker"].notna()]["ticker"].unique())
        existing = None
        print(f"Fetching prices for {len(tickers_to_fetch)} tickers...")

    min_date = holdings_df["date"].min()
    max_date = holdings_df["date"].max()
    start_dt = (datetime.strptime(min_date, "%Y-%m-%d") - timedelta(days=7)).strftime("%Y-%m-%d")
    end_dt = (datetime.strptime(max_date, "%Y-%m-%d") + timedelta(days=3)).strftime("%Y-%m-%d")

    batch_size = 50
    all_price_dfs = []
    for i in range(0, len(tickers_to_fetch), batch_size):
        batch = tickers_to_fetch[i:i + batch_size]
        ticker_str = " ".join(batch)
        try:
            data = yf.download(ticker_str, start=start_dt, end=end_dt,
                               progress=False, auto_adjust=True)
            if isinstance(data.columns, pd.MultiIndex):
                closes = data["Close"]
            else:
                closes = data[["Close"]].rename(columns={"Close": batch[0]})
            all_price_dfs.append(closes)
        except Exception as e:
            print(f"  Error fetching batch {i//batch_size}: {e}")
        time.sleep(1)

    if all_price_dfs:
        new_prices = pd.concat(all_price_dfs, axis=1)
        if existing is not None:
            prices = pd.concat([existing, new_prices], axis=1)
        else:
            prices = new_prices
    elif existing is not None:
        prices = existing
    else:
        prices = pd.DataFrame()

    if not prices.empty:
        prices.index = pd.to_datetime(prices.index)
        prices = prices.sort_index()
        prices.to_csv(PRICES_CACHE_PATH)
        print(f"Prices cached: {len(prices.columns)} tickers, {len(prices)} days")

    return prices


def build_portfolio_snapshot(date_str: str, holdings_df: pd.DataFrame,
                             prices_df: pd.DataFrame) -> pd.DataFrame:
    """특정 날짜의 보유종목 + 종가 매핑 스냅샷."""
    day_holdings = holdings_df[holdings_df["date"] == date_str].copy()
    if day_holdings.empty:
        return pd.DataFrame()

    price_date = pd.Timestamp(date_str)
    if price_date in prices_df.index:
        price_row = prices_df.loc[price_date]
    else:
        mask = prices_df.index <= price_date
        if mask.any():
            price_row = prices_df.loc[mask].iloc[-1]
        else:
            price_row = pd.Series(dtype=float)

    day_holdings["close_usd"] = day_holdings["ticker"].map(
        lambda t: price_row.get(t) if pd.notna(t) and t in price_row.index else None
    )

    return day_holdings.sort_values("weight_pct", ascending=False).reset_index(drop=True)


def compare_dates(date1: str, date2: str, holdings_df: pd.DataFrame) -> pd.DataFrame:
    """두 날짜의 보유종목 비중 변화 비교."""
    h1 = holdings_df[holdings_df["date"] == date1][["ticker", "name", "weight_pct"]].copy()
    h2 = holdings_df[holdings_df["date"] == date2][["ticker", "name", "weight_pct"]].copy()
    h1 = h1.rename(columns={"weight_pct": f"weight_{date1}"})
    h2 = h2.rename(columns={"weight_pct": f"weight_{date2}", "name": "name_2"})

    merged = pd.merge(h1, h2, on="ticker", how="outer")
    merged["name"] = merged["name"].fillna(merged["name_2"])
    merged = merged.drop(columns=["name_2"], errors="ignore")

    col1 = f"weight_{date1}"
    col2 = f"weight_{date2}"
    merged[col1] = merged[col1].fillna(0)
    merged[col2] = merged[col2].fillna(0)
    merged["change"] = merged[col2] - merged[col1]

    merged["status"] = "hold"
    merged.loc[merged[col1] == 0, "status"] = "NEW"
    merged.loc[merged[col2] == 0, "status"] = "EXIT"

    return merged.sort_values("change", ascending=False).reset_index(drop=True)


def weight_timeseries(ticker: str, holdings_df: pd.DataFrame) -> pd.DataFrame:
    """특정 종목의 비중 시계열."""
    mask = holdings_df["ticker"] == ticker
    ts = holdings_df.loc[mask, ["date", "weight_pct", "name"]].copy()
    ts["date"] = pd.to_datetime(ts["date"])
    return ts.sort_values("date").reset_index(drop=True)


def top_holdings_over_time(holdings_df: pd.DataFrame, n: int = 10) -> pd.DataFrame:
    """날짜별 상위 N종목 비중 피벗 테이블."""
    equity = holdings_df[holdings_df["ticker"].notna() & ~holdings_df["is_futures"]].copy()
    equity["rank"] = equity.groupby("date")["weight_pct"].rank(ascending=False, method="first")
    top = equity[equity["rank"] <= n].copy()
    pivot = top.pivot_table(index="date", columns="ticker", values="weight_pct", fill_value=0)
    pivot.index = pd.to_datetime(pivot.index)
    return pivot.sort_index()


def summary_stats(holdings_df: pd.DataFrame) -> dict:
    """전체 데이터셋 요약 통계."""
    dates = sorted(holdings_df["date"].unique())
    all_tickers = holdings_df[holdings_df["ticker"].notna()]["ticker"].unique()
    return {
        "total_dates": len(dates),
        "date_range": f"{dates[0]} ~ {dates[-1]}",
        "unique_tickers": len(all_tickers),
        "avg_holdings_per_day": holdings_df.groupby("date").size().mean(),
        "avg_cash_pct": holdings_df[holdings_df["is_cash"]]["weight_pct"].mean(),
        "most_frequent_tickers": (
            holdings_df[holdings_df["ticker"].notna()]
            .groupby("ticker").size()
            .sort_values(ascending=False)
            .head(20)
            .to_dict()
        ),
    }


if __name__ == "__main__":
    print("=" * 60)
    print("TIME 미국나스닥100 액티브 ETF (426030) 보유종목 분석")
    print("=" * 60)

    print("\n[1/3] 보유종목 스캔...")
    holdings = scan_all_holdings()

    if holdings.empty:
        print("No data found.")
        exit(1)

    print("\n[2/3] 종가 다운로드...")
    prices = fetch_all_prices(holdings)

    print("\n[3/3] 분석 결과 생성...")
    stats = summary_stats(holdings)
    print(f"\n총 {stats['total_dates']}일 데이터 ({stats['date_range']})")
    print(f"고유 종목 수: {stats['unique_tickers']}개")
    print(f"일평균 보유종목: {stats['avg_holdings_per_day']:.1f}개")
    print(f"평균 현금 비중: {stats['avg_cash_pct']:.2f}%")

    print("\n가장 자주 보유한 종목 TOP 20:")
    for ticker, count in stats["most_frequent_tickers"].items():
        pct = count / stats["total_dates"] * 100
        print(f"  {ticker:8s} — {count:4d}일 ({pct:.0f}%)")

    latest_date = sorted(holdings["date"].unique())[-1]
    snapshot = build_portfolio_snapshot(latest_date, holdings, prices)
    if not snapshot.empty:
        print(f"\n최신 포트폴리오 ({latest_date}):")
        cols = ["ticker", "name", "weight_pct", "close_usd", "quantity"]
        print(snapshot[cols].head(15).to_string(index=False))

    holdings.to_csv(HOLDINGS_CACHE_PATH, index=False)
    print(f"\n결과 저장: {HOLDINGS_CACHE_PATH}")
    print("Done.")
